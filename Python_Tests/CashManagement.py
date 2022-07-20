from openpyxl import load_workbook
import decimal
import datetime
from datetime import datetime, date, timedelta

def TxtExl():
    TxtExl_StmtHdrs()
    TxtExl_StmtLns()
    TxtExl_StmtBls()

#Statement Headers Sheet
def TxtExl_StmtHdrs():

  #Abre un archivo txt, r = Read
  file = open('src/hsbc.txt', 'r')
  f = file.readlines()

  # Abre un Excel YA creado.
  file_path = 'src/CashManagementBankStatementImportTemplate.xlsm'
  wb = load_workbook(file_path, read_only=False, keep_vba=True)
  ws = wb['Statement Headers']  # or wb.active

  #Guarda contenido del Txt en un List

  newList = []
  for line in f:
    if line[-1] == '\n':
      newList.append(line)
    else:
      newList.append(line)  

  #Iteración en la newList e Inserción a Excel

  for index, item in enumerate(newList):

    # Variables
    column = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    #Identifica el tipo de Tag
    #"Switch"
    #Caso Tag 20
    if(item.startswith(":20:")):
      row = 4
      col = 1
      colId = 0
      tag = item
      index = [4, 14]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      # Inserción de cada Split del Item por Celda predefinida
      for id, items in enumerate(parts):
          while ws.cell(row, col).value:
            row+=1
          ws[column[colId] + str(row)] = items
          colId += 1
          col += 1
    #Caso Tag 25 / Inserta Fecha
    elif(item.startswith(":25:")):
      mesActual = datetime.now().month
      fecha = Mes_Dia_Final(date(2022, mesActual, 1))
      row = 4
      col = 2
      tag = item.split(":25:", 1)
      for items in enumerate(tag):
          while ws.cell(row, col).value:
            row+=1
          ws["B" + str(row)] = items[1]
          ws["D" + str(row)] = fecha
          col += 1
    #Caso Tag 28C
    elif(item.startswith(":28C:")):
      row = 1
      col = 4
      colId = 3
      tag = item.split(":28C:", 1)
      aux = tag[1].split("/")
      for id, items in enumerate(aux):
          while ws.cell(row, col).value:
            row+=1
          ws[column[colId] + str(row)] = items
          colId += 1
          col += 1
    #Caso Tag 60F
    elif(item.startswith(":60F:")):
      row = 4
      col = 5
      tag = item
      index = [5, 6, 12, 15, 30]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      #for items in enumerate(parts):
      while ws.cell(row, col).value:
        row+=1
      ws[column[4] + str(row)] = parts[2]
          #col += 1
    #Caso Tag 61 // Arreglar el salto de Línea
    elif(item.startswith(":61:")):
      row = 1
      col = 11
      colId = 10
      tag = item
      index = [4, 10, 14, 15, 30, 34, 70]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      for id, items in enumerate(parts):
         while ws.cell(row, col).value:
           row+=1
         ws[column[colId] + str(row)] = items
         colId += 1
         col += 1
    #Caso Tag 86
    #elif(item.startswith(":86:")):
    #  row = 1
    #  col = 17
    #  tag = item.split(":86:", 1)
    #  for items in enumerate(tag):
    #      while ws.cell(row, col).value:
    #        row+=1
    #      ws["Q" + str(row)] = items[1]
    #      col += 1
    #Caso Tag 62M
    elif(item.startswith(":62M:")):
      row = 1
      col = 18
      colId = 17
      tag = item
      index = [5, 6, 12, 15, 30]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      for id, items in enumerate(parts):
         while ws.cell(row, col).value:
           row+=1
         ws[column[colId] + str(row)] = items
         colId += 1
         col += 1
    else:
      item = ""

  print("Statement Headers Updated")
  
  wb.save(file_path)
#Statement  Balance Sheet
def TxtExl_StmtBls():

  #Abre un archivo txt, r = Read
  file = open('src/hsbc.txt', 'r')
  f = file.readlines()

  # Abre un Excel YA creado.
  file_path = 'src/CashManagementBankStatementImportTemplate.xlsm'
  wb = load_workbook(file_path, read_only=False, keep_vba=True)
  ws = wb['Statement Balances']  # or wb.active

  #Guarda contenido del Txt en un List

  newList = []
  for line in f:
    if line[-1] == '\n':
      newList.append(line)
    else:
      newList.append(line)  

  #Iteración en la newList e Inserción a Excel

  for index, item in enumerate(newList):

    # Variables
    column = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    #Identifica el tipo de Tag
    #"Switch"
    #Caso Tag 20
    if(item.startswith(":20:")):
      row = 4
      col = 1
      colId = 0
      tag = item
      index = [4, 14]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      # Inserción de cada Split del Item por Celda predefinida
      for id, items in enumerate(parts):
          while ws.cell(row, col).value:
            row+=1
          ws[column[colId] + str(row)] = items
          colId += 1
          col += 1
    #Caso Tag 25
    elif(item.startswith(":25:")):
      row = 4
      col = 2
      tag = item.split(":25:", 1)
      for items in enumerate(tag):
          while ws.cell(row, col).value:
            row+=1
          ws[column[1] + str(row)] = items[1]
          ws[column[5] + str(row)] = "CRDT"
          col += 1
    #Caso Tag 28C
    elif(item.startswith(":28C:")):
      row = 1
      col = 4
      colId = 3
      tag = item.split(":28C:", 1)
      aux = tag[1].split("/")
      for id, items in enumerate(aux):
          while ws.cell(row, col).value:
            row+=1
          ws[column[colId] + str(row)] = items
          colId += 1
          col += 1
    #Caso Tag 60F
    elif(item.startswith(":60F:")):
      row = 4
      col = 5
      tag = item
      index = [5, 6, 12, 15, 30]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      #for items in enumerate(parts):
      while ws.cell(row, col).value:
        row+=1
      ws[column[4] + str(row)] = parts[2]
      col += 1
    #Caso Tag 61 // Arreglar el salto de Línea
    elif(item.startswith(":61:")):
      row = 1
      col = 11
      colId = 10
      tag = item
      index = [4, 10, 14, 15, 30, 34, 70]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      for id, items in enumerate(parts):
         while ws.cell(row, col).value:
           row+=1
         ws[column[colId] + str(row)] = items
         colId += 1
         col += 1
    #Caso Tag 86
    #elif(item.startswith(":86:")):
    #  row = 1
    #  col = 17
    #  tag = item.split(":86:", 1)
    #  for items in enumerate(tag):
    #      while ws.cell(row, col).value:
    #        row+=1
    #      ws["Q" + str(row)] = items[1]
    #      col += 1
    #Caso Tag 62M
    elif(item.startswith(":62M:")):
      row = 1
      col = 18
      colId = 17
      tag = item
      index = [5, 6, 12, 15, 30]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      for id, items in enumerate(parts):
         while ws.cell(row, col).value:
           row+=1
         ws[column[colId] + str(row)] = items
         colId += 1
         col += 1
    else:
      item = ""

  print("Statement Balance Updated")
  
  wb.save(file_path)
#Statement Lines Sheet
def TxtExl_StmtLns():

  #Abre un archivo txt, r = Read
  file = open('src/hsbc.txt', 'r')
  f = file.readlines()

  # Abre un Excel YA creado.
  file_path = 'src/CashManagementBankStatementImportTemplate.xlsm'
  wb = load_workbook(file_path, read_only=False, keep_vba=True)
  ws = wb['Statement Lines']  # or wb.active

  #Guarda contenido del Txt en un List

  newList = []
  for line in f:
    if line[-1] == '\n':
      newList.append(line)
    else:
      newList.append(line)  

  #Iteración en la newList e Inserción a Excel

  for index, item in enumerate(newList):

    # Variables
    column = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    #Identifica el tipo de Tag
    #"Switch"
    #Caso Tag 20
    if(item.startswith(":20:")):
      row = 4
      col = 1
      colId = 0
      tag = item
      index = [4, 14]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      # Inserción de cada Split del Item por Celda predefinida
      for id, items in enumerate(parts):
          while ws.cell(row, col).value:
            row+=1
          ws[column[colId] + str(row)] = items
          colId += 1
          col += 1
    #Caso Tag 25 / #Account & Line Number & CRDT
    elif(item.startswith(":25:")):
      row = 4
      col = 2
      ln_number = 0
      tag = item.split(":25:", 1)
      #for items in enumerate(tag):
      while ws.cell(row, col).value:
        row+=1
        ln_number += 1
      ws[column[1] + str(row)] = tag[1]
      ws[column[2] + str(row)] = ln_number
      col += 1  
    #Caso Tag 28C
    elif(item.startswith(":28C:")):
      row = 1
      col = 4
      colId = 3
      tag = item.split(":28C:", 1)
      aux = tag[1].split("/")
      for id, items in enumerate(aux):
          while ws.cell(row, col).value:
            row+=1
          ws[column[colId] + str(row)] = items
          colId += 1
          col += 1
    #Caso Tag 60F
    elif(item.startswith(":60F:")):
      row = 4
      col = 6
      tag = item
      index = [5, 6, 12, 15, 30]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      moneda = parts[2]
      for items in enumerate(parts):
          while ws.cell(row, col).value:
            row+=1
          #ws["E" + str(row)] = moneda
          col += 1
    #Caso Tag 61 // Arreglar el salto de Línea
    elif(item.startswith(":61:")):
      row = 1
      col = 11
      colId = 10
      tag = item
      index = [4, 10, 14, 15, 30, 34, 70]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      for id, items in enumerate(parts):
         while ws.cell(row, col).value:
           row+=1
         ws[column[colId] + str(row)] = items
         colId += 1
         col += 1
    #Caso Tag 86
    elif(item.startswith(":86:")):
      row = 4
      col = 5
      tag = item.split(":86:", 1)
      for items in enumerate(tag):
          while ws.cell(row, col).value:
            row+=1
          ws[column[4] + str(row)] = items[1]
          col += 1
    #Caso Tag 62M
    elif(item.startswith(":62M:")):
      row = 1
      col = 18
      colId = 17
      tag = item
      index = [5, 6, 12, 15, 30]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      for id, items in enumerate(parts):
         while ws.cell(row, col).value:
           row+=1
         ws[column[colId] + str(row)] = items
         colId += 1
         col += 1
    else:
      item = ""

  print("Statement Lines Updated")
  
  wb.save(file_path)

def Mes_Dia_Final(any_day):
    # The day 28 exists in every month. 4 days later, it's always next month
    next_month = any_day.replace(day=28) + timedelta(days=4)
    # subtracting the number of the current day brings us back one month
    return next_month - timedelta(days=next_month.day)