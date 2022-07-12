from openpyxl import load_workbook
import decimal

def TxtExl():

  #Abre un archivo txt, r = Read
  file = open('src/hsbc.txt', 'r')
  f = file.readlines()

  # Abre un Excel YA creado
  file_path = 'src/Prueba.xlsx'
  wb = load_workbook(file_path)
  ws = wb['Hoja1']  # or wb.active

  #Guarda contenido del Txt en un List

  newList = []
  for line in f:
    if line[-1] == '\n':
      newList.append(line)
    else:
      newList.append(line)  

  #Iteración en la newList e Inserción a Excel

  for index, item in enumerate(newList):

    # Variables
    column = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    #Identifica el tipo de Tag
    #"Switch"
    #Caso Tag 20
    if(item.startswith(":20:")):
      row = 1
      col = 1
      colId = 0
      tag = item
      index = [4, 14]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      # Inserción de cada Split del Item por Celda predefinida
      for id, items in enumerate(parts):
          while ws.cell(row, col).value:
            row+=1
          ws[column[colId] + str(row)] = items
          colId += 1
          col += 1
    #Caso Tag 25
    elif(item.startswith(":25:")):
      row = 1
      col = 3
      tag = item.split(":25:", 1)
      for items in enumerate(tag):
          while ws.cell(row, col).value:
            row+=1
          ws["C" + str(row)] = items[1]
          col += 1
    #Caso Tag 28C
    elif(item.startswith(":28C:")):
      row = 1
      col = 4
      colId = 3
      tag = item.split(":28C:", 1)
      aux = tag[1].split("/")
      for id, items in enumerate(aux):
          while ws.cell(row, col).value:
            row+=1
          ws[column[colId] + str(row)] = items
          colId += 1
          col += 1
    #Caso Tag 60F
    elif(item.startswith(":60F:")):
      row = 1
      col = 6
      colId = 5
      tag = item
      index = [5, 6, 12, 15, 30]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      for id, items in enumerate(parts):
         while ws.cell(row, col).value:
           row+=1
         ws[column[colId] + str(row)] = items
         colId += 1
         col += 1
    #Caso Tag 61 // Arreglar el salto de Línea
    elif(item.startswith(":61:")):
      row = 1
      col = 11
      colId = 10
      tag = item
      index = [4, 10, 14, 15, 30, 34, 70]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      for id, items in enumerate(parts):
         while ws.cell(row, col).value:
           row+=1
         ws[column[colId] + str(row)] = items
         colId += 1
         col += 1
    #Caso Tag 86
    elif(item.startswith(":86:")):
      row = 1
      col = 17
      tag = item.split(":86:", 1)
      for items in enumerate(tag):
          while ws.cell(row, col).value:
            row+=1
          ws["Q" + str(row)] = items[1]
          col += 1
    #Caso Tag 62M
    elif(item.startswith(":62M:")):
      row = 1
      col = 18
      colId = 17
      tag = item
      index = [5, 6, 12, 15, 30]
      parts = [tag[i:j] for i,j in zip(index, index[1:] + [None])]
      for id, items in enumerate(parts):
         while ws.cell(row, col).value:
           row+=1
         ws[column[colId] + str(row)] = items
         colId += 1
         col += 1
    else:
      item = ""

  print("Excel Created")
  
  wb.save(file_path)
