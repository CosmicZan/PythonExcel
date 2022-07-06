from openpyxl import load_workbook

def Test():
  file_path = 'src/Prueba.xlsx'

  wb = load_workbook(file_path)

  ws = wb['Hoja1']  # or wb.active

  ws['G6'] = 123
  ws['G7'] = 123

  wb.save(file_path)